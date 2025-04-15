import unicodedata
import traceback
import csv
import os
import json
import subprocess
import sys
#pdfminer.six
import shlex
import re

from types import SimpleNamespace

from .model_data import get_model_data
from .model_data import load_model_data
from .prompt_data import get_prompt_data
from .prompt_data import load_prompt_data
from .parse_pubmed_json import parse_pubmed_data

from .claude_engine import ClaudeEngine
from .open_ai_engine import OpenAIEngine
from .external_engine import ExternalEngine
 

def is_one_token(s):
    return bool(re.fullmatch(r"\[\w+\]", s))

DEFAULT_MAX_PROMPT_LENGTH = 100000
DEFAULT_MAX_TOKENS = 4096

self_data =  SimpleNamespace()
model_data = {}

print("starting...")


def isYes(answer,param):
    text = answer.lower().strip()
    
    return text == "yes" or text == "yes." or text == "y" or text == "true" or text == "t" or text == "1"
def isNo(answer,param):
    text = answer.lower().strip()
    return text == "no" or text == "no." or text == "n" or text == "false" or text == "f" or text == "0"
    

def isShort(answer,param):
    length = 5
    try:
        length = int(param)
    
    except:
        return False
    return len(answer,param) < length
def isLong(answer,param):
    length = 5
    try:
        length = int(param)
    
    except:
        return False
    return len(answer,param) >= length
    
def isGreaterThan(answer,param):
    try:
        return float(answer) > float(param)
    except:
        return False
def isLessThan(answer,param):
    try:
        return float(answer) < float(param)
    except:
        return False
    
    
def isNumber(answer,param):
    try:
        float(answer)
        return True
    except:
        return False
def isJson(answer,param):
    try:
        json.loads(answer)
        return True
    except:
        return False
def isNotNumber(answer,param):
    return not isNumber(answer,param)
def isNotJson(answer,param):
    return not isJson(answer,param)

def isJsonList(answer,param):
    
    
    try:
        data = json.loads(answer)
        answer =  isinstance(data, list)  
    
        return answer
    except:
    
        return False
def isNotJsonList(answer,param):
    return not isJsonList(answer,param)

def isCommaSeparatedList(answer,param):
    return "," in answer
def isNotCommaSeparatedList(answer,param):
    return not isCommaSeparatedList(answer,param)

def isError(answer,param):
    return self_data.script_returncode != 0


prechecks = {"is_yes": isYes, "is_no": isNo, "is_number": isNumber, "is_json": isJson,isNotNumber: isNotNumber, isNotJson: isNotJson, "is_comma_separated_list": isCommaSeparatedList, "is_not_comma_separated_list": isNotCommaSeparatedList, "is_json_list": isJsonList, "is_not_json_list": isNotJsonList,
"yes": isYes, "no": isNo, "number": isNumber, "json": isJson,isNotNumber: isNotNumber, isNotJson: isNotJson, "comma_separated_list": isCommaSeparatedList, "not_comma_separated_list": isNotCommaSeparatedList, "json_list": isJsonList, "not_json_list": isNotJsonList,
"is_short": isShort,"is_long" : isLong, "is_greater_than": isGreaterThan, "is_less_than": isLessThan, "is_script_error": isError

}





def setup_data():
    global model_data
    model_data = get_model_data()
    print(model_data)
    
    self_data.start_from = model_data.get("start_from",0)
    self_data.max_docs = model_data.get("max_documents",sys.maxsize)

    self_data.precheck_system = "You are helping to automate a workflow.  You will be asked a question to verify the information you have given. You must only answer the question in EXACTLY the format requested. The answer will only ever be read by a computer, NEVER add any other commentary or automated processing will fail. It is more important to give a correctly formatted answer than to be sure the answer is correct."
    self_data.precheck_system = model_data.get("precheck_system",self_data.precheck_system)
    self_data.max_tokens = model_data.get("max_tokens",DEFAULT_MAX_TOKENS)
    
    self_data.model_data = model_data

    self_data.which_api = model_data["model"]

    self_data.api_key = model_data.get("api_key")
    self_data.data_folder = model_data.get("files_folder","files")
    if self_data.data_folder == []:
        self_data.data_folder = "."

    self_data.data_folder = resolve_path(self_data.data_folder)

    self_data.max_prompt_length = model_data.get("max_prompt_length",DEFAULT_MAX_PROMPT_LENGTH)
    self_data.max_prompt_length = int(self_data.max_prompt_length)
    self_data.max_doc_length = model_data.get("max_document_length",sys.maxsize)
    self_data.max_doc_length = int(self_data.max_doc_length)




    self_data.column_name = "pubmed_id"



    # There is an alternative to use a local script to get pubmed data
    self_data.use_pubmed_api = True
    if "use_pubmed_api" in model_data:
        if model_data.get("use_pubmed_api","false").strip().lower() == "false":
            self_data.use_pubmed_api = False

    self_data.use_pubmed_search = False

    if "use_pubmed_search" in model_data:
        if model_data.get("use_pubmed_search","").strip().lower() == "true":
            self_data.use_pubmed_search = True


    # Responses are stored so that they are not repeated later
    # If you want to clear the cache, delete the cache folder, or you can change the key in the specific api file

    self_data.data_cache_folder = model_data.get("self_data.data_cache_folder", "cache/data")
    self_data.cache_folder = model_data.get("cache_folder", "cache/api")



    self_data.data_store = {}
    self_data.output_data = {}
    self_data.final_output = {}
    self_data.debug = {}

    self_data.ordered_column_list = []
    self_data.script_returncode = 0





def setup():
    
    setup_data()
    
    os.makedirs(self_data.data_cache_folder, exist_ok=True)
    os.makedirs(self_data.cache_folder, exist_ok=True)
    
    


    if self_data.which_api == "claude":
        self_data.llm_engine = ClaudeEngine(
        key=self_data.api_key, cache_folder=self_data.cache_folder, max_tokens = self_data.max_tokens)
    if self_data.which_api == "openai":
        self_data.llm_engine = OpenAIEngine(key=self_data.api_key, cache_folder=self_data.cache_folder, max_tokens = self_data.max_tokens)
    if self_data.which_api == "external":
        self_data.llm_engine = ExternalEngine(cache_folder=self_data.cache_folder, max_tokens = self_data.max_tokens)
    
    
        
 






def preprocess_prompt_old(prompt,escape = False):
    
    
    for key in self_data.data_store:
        new_text = self_data.data_store[key]
        if escape:
            new_text = repr(new_text)

        prompt = prompt.replace(f"[{key}]", new_text)

    return prompt
    
def preprocess_prompt(prompt, max_length=None, escape=False, overlap=500):
 
    if max_length is None:
        max_length = self_data.max_prompt_length
    # Track original text and substitutions for splitting later if needed
    substitutions = []
    result = prompt
    
    # First pass: identify all substitutions
    for key in self_data.data_store:
        placeholder = f"[{key}]"
        if placeholder in result:
            new_text = self_data.data_store[key]
            if escape:
                new_text = repr(new_text)
            
            # Record information about each substitution
            for match_pos in find_all_occurrences(result, placeholder):
                substitutions.append({
                    'key': key,
                    'placeholder': placeholder,
                    'replacement': new_text,
                    'position': match_pos,
                    'length': len(new_text)
                })
    
    # Sort substitutions by position (to maintain order)
    substitutions.sort(key=lambda x: x['position'])
    
    # Apply substitutions to get the processed text
    processed_prompt = prompt
    for sub in substitutions:
        processed_prompt = processed_prompt.replace(sub['placeholder'], sub['replacement'], 1)
    
    # If the processed prompt is within limits, return it as a single-item list
    if len(processed_prompt) <= max_length:
        return [processed_prompt]
    
    # Otherwise, find the longest substitution to split
    if not substitutions:
        # If no substitutions but still too long, simply truncate
        return [processed_prompt[:max_length]]
    
    # Find the longest substitution
    longest_sub = max(substitutions, key=lambda x: x['length'])
    
    if longest_sub['length'] <= overlap:
        # If even the longest substitution is too short to split meaningfully,
        # truncate the result
        return [processed_prompt[:max_length]]
    
    # Split the longest substitution
    original_text = longest_sub['replacement']
    split_point = len(original_text) // 2
    
    # Ensure the split point provides proper overlap
    part1 = original_text[:split_point + overlap]
    part2 = original_text[split_point:]
    
    # Create two new data stores with the split substitution
    new_prompts = []
    
    # For each split part, create a new prompt
    for i, part in enumerate([part1, part2]):
        # Create a copy of the data store with the modified substitution
        temp_data_store = dict(self_data.data_store)
        temp_data_store[longest_sub['key']] = part
        
        # Create a temporary data context
        original_data_store = self_data.data_store
        self_data.data_store = temp_data_store
        
        # Recursively process with the updated substitution
        split_results = preprocess_prompt(prompt, max_length, escape, overlap)
        new_prompts.extend(split_results)
        
        # Restore original data store
        self_data.data_store = original_data_store
    
    return new_prompts

def find_all_occurrences(text, substring):
    """Helper function to find all occurrences of a substring in text"""
    positions = []
    start = 0
    while True:
        start = text.find(substring, start)
        if start == -1:
            break
        positions.append(start)
        start += 1
    return positions

def get_text_from_prompt(prompt, system):
    

    if prompt.startswith("#py"):
        prompt = prompt[3:]
        if prompt.startswith("#python"):
            prompt = prompt[7:]

        full_prompt = preprocess_prompt(prompt,escape=True, max_length = sys.maxsize)
        result = prompt
        full_prompt = full_prompt[0]
    

        try:
            result = str(eval(full_prompt))
        except Exception as e:
            print("Error processing python prompt")
            print(e)
            traceback.print_exc(file=sys.stdout)

        result = ' '.join(result.split())
        return result
    
    
    #special case to indicate that the prompt should be generated but not processed
#i.e., for retrieving variables or direct quotes from the input file
    if prompt.startswith("#"):
        full_prompt = preprocess_prompt(prompt,max_length = sys.maxsize)
        full_prompt = full_prompt[0]
        #more special case to call a script
        if prompt.startswith("#!"):
            full_prompt = full_prompt[0]
            params = full_prompt[2:].split(" ",1)

            script_folder = model_data.get("script_folder", "scripts")

            exe = os.path.join( script_folder, params[0] )
            params = params[1]
            result = subprocess.run([exe,params], shell=True, text=True, capture_output=True)
            if result.returncode != 0:
                self_data.script_returncode  = result.returncode
                
            
            result = result.stdout
 
        else:
            result = full_prompt[1:]
    else:
        cache_prompt = None

        #special tag to splt into two prompts, first part is requested to be cached
        if "[!cache]" in prompt:
            cache_prompt_pre, prompt = prompt.split("[!cache]",1)

            cache_prompt = preprocess_prompt(cache_prompt_pre)
            #If it's too big and was split, we won't cache it

            if len(cache_prompt) > 1:
                cache_prompt = None
                prompt = cache_prompt_pre +" " + prompt
            else:
                cache_prompt = cache_prompt[0]
                


        full_prompt = preprocess_prompt(prompt)


        results = []
        for pr in full_prompt:
            if cache_prompt:
                results.append(self_data.llm_engine.promptWithCache(pr, cache_prompt, system))
            else:
                results.append(self_data.llm_engine.prompt(pr, system))
        result = " ".join(results)


    #remove characters that are not printable, including newlines and tabs
    result = ' '.join(result.split())

    return result

def process_line(line):
    system = line["system"]

    name = line["name"]
    if len(name) == 0 and len(line["prompts"]) == 0:
        return self_data.data_store["reply"]
    
    preCheck = None
    result = None
    if line["skipTest"]:
        preCheck = line["skipPrompt"]
        
        preCheckResult = get_text_from_prompt(preCheck, self_data.precheck_system)
        preCheckTest = line["skipTest"].split()
        param = preCheckTest[1] if len(preCheckTest) > 1 else ""
    
        preCheckTest = preCheckTest[0]
        
        preCheckTestFunction = isYes
        if preCheckTest in prechecks:
            preCheckTestFunction = prechecks[preCheckTest]

        #if the answer is yes, then we jump to the next stage
        #If no, we will use this prompt
        
        if preCheckTestFunction(preCheckResult,param):
        #    print("Skipping",preCheckTest)
            return self_data.data_store["reply"]
 
    for prompt in line["prompts"]:
        result = get_text_from_prompt(prompt, system)


        if result.lower() == "!cancel!":
            print("cancelled")
            return None

        self_data.data_store["reply"] = result
        self_data.reply_count += 1
        self_data.data_store[f"reply_{self_data.reply_count}"] = result



    if result:
        self_data.data_store["reply"] = result
        self_data.data_store[name] = result

        
        if line["dataOut"]:
             
            self_data.output_data[name] = result


            if name not in self_data.ordered_column_list:
                self_data.ordered_column_list.append(name)
    else:
        print("No result for",name)

    #    print(f"{name}: {result}")
    return result


def process_document(pmid,document_data):

    try:
        text = document_data["text"]
        sections = document_data["sections"]

        result = None
        self_data.reply_count = 0
        self_data.output_data = {}


        self_data.data_store = {"paper": text}
        self_data.data_store["cancel"] = "!cancel!"
        


        for section in sections:
            self_data.data_store[section] = sections[section]

        for m in model_data:
            if m.startswith("["):
                variable = m[1:-1]
                self_data.data_store[variable] = model_data[m]


        print(f"Processing {pmid}")
        prompt_data = get_prompt_data()
        
        for process in prompt_data:

            result = process_line(process)

            if result is None:
                break

        if result is not None:
            result = self_data.output_data.copy()
        self_data.debug[pmid] = self_data.data_store.copy()

        return result

    except Exception as e:
        print("Error processing document")
        print("error", e)

        traceback.print_exc(file=sys.stdout)
        return None
    
def get_pubmed_from_local(pubmed_id):
    script_path = model_data["get_pubmed_path"]

    #call an external script to get the data, passing in the pubmed id
    data = []
    try:
        data = subprocess.check_output([script_path, pubmed_id])
        data = json.loads(data)
    except subprocess.CalledProcessError as e:
        print("Error",e)

    return data

def get_text_from_local(filename):


    filename = os.path.join(self_data.data_folder,filename)
    
    

    
    if filename.lower().endswith(".pdf"):
        try:
            from pdfminer.high_level import extract_text
            all_text = extract_text(filename)

        except Exception as e:
            print("Error processing pdf - pfminder.six must be installed, or use text or json files")
            print(e)
            traceback.print_exc(file=sys.stdout)     
            
    elif filename.lower().endswith(".json"):
        with open(filename, 'r', encoding='utf-8') as file:
            data = json.load(file)
            all_text = ""
            if "text" in data:
                all_text = data["text"]
            elif "sections" in data:
                for section in data["sections"]:
                    all_text += data["sections"][section] + "\n"

                    
            if "sections" in data:
                data = {"text": all_text, "sections" : data["sections"]}
                return data

    else:
        with open(filename, 'r', encoding='utf-8', errors='ignore') as file:

            print("read",filename)
            all_text = file.read()



    all_text = unicodedata.normalize("NFKC", all_text)



    data = {"text": all_text, "sections" : {"paper":all_text}}

    return data

def get_pubmed_from_api(pubmed_id):
    
    try:
       import requests
    except:
       raise RuntimeError("To use an External PubMed API, requests must be installed.")
    api_url = "https://www.ncbi.nlm.nih.gov/research/bionlp/RESTful/pmcoa.cgi/BioC_json/" + str(pubmed_id) + "/unicode"
    if "pubmed_url" in model_data:
        api_url = model_data["pubmed_url"] + str(pubmed_id) + "/unicode"


    data = []
    try:
        response = requests.get(api_url)
        response.raise_for_status()  # Raise an exception for 4XX/5XX responses
        data = response.json()

    except requests.exceptions.HTTPError as http_err:
        print(f"HTTP error occurred: {http_err}")
    except requests.exceptions.RequestException as req_err:
        print(f"Request error occurred: {req_err}")

    return data

# Function to fetch PubMed data via API or from a local file
def fetch_pubmed_data(pubmed_id, sections_to_extract, data_folder):

#if it is numberical or a PMC id, then we assume it is pubmed
#otherwise we assume it is a local file

    is_pubmed = (pubmed_id.isnumeric() or (pubmed_id[:3] == "PMC" and pubmed_id[3:].isnumeric()))
       

    json_file_path = os.path.join(data_folder, f"{pubmed_id}.json")

    # Check if the JSON file already exists, to cache it
    if os.path.exists(json_file_path):
        with open(json_file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)


    else:

    
        if is_pubmed:
                
            # Fetch the data from PubMed API (alternative is a local script)
            if self_data.use_pubmed_api:
                data = get_pubmed_from_api(pubmed_id)
            else:
                data = get_pubmed_from_local(pubmed_id)
        else:
            data = get_text_from_local(pubmed_id)
         
        with open(json_file_path, 'w', encoding='utf-8') as json_file:


            json.dump(data, json_file, ensure_ascii=False, indent=4)

    # Extract the relevant sections from the JSON data
    if is_pubmed:
        parsed_data = parse_pubmed_data(data, sections_to_extract)
    else:
        parsed_data = data
        

    return parsed_data




# Function to process each PubMed ID
def process_pubmed_id(pubmed_id, processed_documents, sections_to_extract, data_folder):
    document_data  = fetch_pubmed_data(pubmed_id, sections_to_extract, data_folder)
    document_text = document_data.get("text")
    
    print("got text",pubmed_id,len(document_text))
    #print(document_text
    if len(document_text) > self_data.max_doc_length:
        print("document too long")
        return
        
    if document_text:
        if len(document_text) > 1:
            processed_documents.append(document_text)
            result = process_document(pubmed_id,document_data)
            if result:
                self_data.final_output[pubmed_id] = result

def normalize_newlines(text):
    if isinstance(text, str):
        text = text[:10000]
        text = text.replace('\"', '')


        text= text.replace('\r', '')
        return text.replace('\n', ' \\n ')
    return text
    
    
def output_csv_old(output_data,outputfile):
    columns = set()
    for key in output_data:
        columns.update(output_data[key].keys())

    # Convert the set to a sorted list to maintain column order
    columns = sorted(columns)
    column_name = self_data.column_name
    # Step 2: Write to CSV
    with open(outputfile, 'w', newline='', encoding='utf-8' ) as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=[column_name] + columns,quoting=csv.QUOTE_ALL)
        writer.writeheader()

        # Step 3: Write data rows
        for key, values in output_data.items():
            row = {column_name: key}  # Start row with the main key
            row.update(values)  # Update the row with the nested dictionary
            normalized_row = {k: normalize_newlines(v) for k, v in row.items()}

            writer.writerow(normalized_row)

 

def output_csv(output_data, outputfile):
    columns = set()
    #column_name is the global for the key of the dictionary
    column_name = self_data.column_name
    # Collect all possible columns from the nested dictionaries
    for key in output_data:
        columns.update(output_data[key].keys())

    # Ensure columns appear in the specified order, filtering only those present in output_data
    ordered_columns = [col for col in self_data.ordered_column_list if col in columns]
    extra_columns = sorted(columns - set(self_data.ordered_column_list))  # Additional columns in sorted order
    final_columns = ordered_columns + extra_columns  # Merge ordered and extra columns

    # Prepend the 'id' column for the key
    header = [column_name] + final_columns

    with open(outputfile, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=header, quoting=csv.QUOTE_ALL)
        writer.writeheader()

        # Write data rows: add the key as the 'id' field for each row
        for key, values in output_data.items():
            row = {column_name: key}
            row.update(values)
            normalized_row = {k: normalize_newlines(v) for k, v in row.items()}
            writer.writerow(normalized_row)

def resolve_path(path):
    """Convert relative paths to absolute, based on working directory."""
    if not os.path.isabs(path):
        return os.path.join(model_data["config_root"], path)
    return path


def read_pubmed_ids(file_path, column_name):
    """
    Reads PubMed IDs from a specified column in either a CSV or XLSX file.
    Returns a list of PubMed IDs.
    """
    pubmed_ids = []

    file_path = resolve_path(file_path)
 
    if file_path.endswith('.csv'):
        with open(file_path, mode='r', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            if column_name not in reader.fieldnames:
                raise ValueError(f"Column '{column_name}' not found in the CSV file.")

            pubmed_ids = [row[column_name] for row in reader if row[column_name]]

    elif file_path.endswith('.xlsx'):
        
        try:
            import openpyxl
        except ModuleNotFoundError:
            raise ValueError("To load Excel files openpyxl must be installed")
  
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        # Find the column index by matching the header
        headers = [cell.value for cell in sheet[1]]
        if column_name not in headers:
            raise ValueError(f"Column '{column_name}' not found in the Excel file.")

        col_index = headers.index(column_name) + 1  # Convert to 1-based index for openpyxl

        # Extract PubMed IDs from the column
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_index, max_col=col_index):
            if row[0].value:
                pubmed_ids.append(str(row[0].value))  # Convert to string to ensure consistency

 
  
    return pubmed_ids





def save_output(data, csv_file, json_file):
    
    
    try:
        output_csv(data, csv_file)
    except Exception as e:
        print("error",csv_file)
        print(e)
        traceback.print_exc(file=sys.stdout)
        
    try:
        with open(json_file, 'w', encoding='utf-8') as json_out:
            json.dump(data, json_out, indent=4)
    except Exception as e:

        print("error",json_file)
        print(e)
        traceback.print_exc(file=sys.stdout)
        
def process_pubmed_ids(pubmed_ids, sections_to_extract, data_folder):
    
    processed_documents = []  # Store processed documents

    output_folder = model_data.get("output_folder", "output")
    output_file = model_data.get("output_file", "output.csv")

    output_file_base = output_file.split(".")[0]
    output_file_ext = output_file.split(".")[-1]

    output_file = os.path.join(output_folder ,output_file) 
    output_file_json = os.path.join(output_folder ,output_file_base+".json") 

    debug_output_file = os.path.join(output_folder ,output_file_base+"_debug."+output_file_ext) 
    debug_output_file_json = os.path.join(output_folder ,output_file_base+"_debug.json")

    os.makedirs(output_folder, exist_ok=True)
    
    
    for pubmed_id in pubmed_ids:
        try:
            process_pubmed_id(pubmed_id, processed_documents, sections_to_extract, data_folder)
        except Exception as e:
            print(e)
            print("error with",pubmed_id)
            traceback.print_exc(file=sys.stdout)
 
        if len(self_data.final_output) > 0:
            save_output(self_data.final_output, output_file, output_file_json)
            save_output(self_data.debug, debug_output_file, debug_output_file_json)
        else:
            print("no output",pubmed_id)
        
    print("Final Output")
    print(self_data.final_output)
    print(self_data.ordered_column_list)




    # Save outputs
    save_output(self_data.final_output, output_file, output_file_json)
    save_output(self_data.debug, debug_output_file, debug_output_file_json)

    return processed_documents

def search_for_pubmed_ids(search_script, search_term,search_options):
    
    pubmed_ids = []

    try:
        
        search_term = re.sub(r'([()])', r' \1 ', search_term).strip()

    
        search_term = search_term + " "+search_options
            #search_term = base64.b64encode(search_term.encode()).decode()

        args = shlex.split(search_term)
        result = subprocess.run([search_script]+args, capture_output=True, text=True, check=True)
        #result = subprocess.run([search_script, "-b", search_term], capture_output=True, text=True, check=True)
        result_text = result.stdout
        
        try:
            #see if it is json
            pubmed_ids = json.loads(result_text)
            pubmed_ids = ["pmid"]+pubmed_ids
        except:
            pubmed_ids = result_text.splitlines()  
            
    except subprocess.CalledProcessError as e:
        print("Error running search script")
        print(e)

    return pubmed_ids[1:],pubmed_ids[0]

def parse_papers(config_file):
 

    load_model_data(config_file)   
    
 
    setup()
    
    load_prompt_data()
    if self_data.use_pubmed_search:
        search_script = model_data.get("pubmed_search_script")
        search_term = model_data.get("pubmed_search_term")
        search_options = model_data.get("pubmed_search_options","")

        pubmed_ids, self_data.column_name = search_for_pubmed_ids(search_script, search_term,search_options)
        
        
    else:
        file_path = model_data.get("documents_data")
        self_data.column_name = model_data.get("column_name")
        pubmed_ids = read_pubmed_ids(file_path, self_data.column_name)
 

    pubmed_ids = pubmed_ids[self_data.start_from:self_data.start_from+self_data.max_docs]
        
    # Get the list of processed documents
    sections_to_extract = model_data.get("sections")
    num_pubmed_ids = len(pubmed_ids)

    print(f"Processing {num_pubmed_ids} documents.")

    processed_documents = process_pubmed_ids(pubmed_ids, sections_to_extract, self_data.data_cache_folder)
    print(f"Processed {len(processed_documents)} documents.")
