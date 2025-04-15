from collections.abc import Sequence, Iterator
from typing import Any, TypeAlias, Literal, TypeGuard
# from typing_extensions import TypeIs

import openpyxl as op
import pandas as pd
import polars as pl

from ka_uts_log.log import Log
from ka_uts_dic.dopddf import DoPdDf
from ka_uts_dic.dopldf import DoPlDf
from ka_uts_dfr.pddf import PdDf
from ka_uts_dfr.pldf import PlDf
from ka_uts_obj.path import Path
from ka_uts_xls.wbop import WbOp
from ka_uts_xls.wsop import WsOp

TyWbOp: TypeAlias = op.workbook.workbook.Workbook
TyWsOp: TypeAlias = op.worksheet.worksheet.Worksheet
TyPdDf: TypeAlias = pd.DataFrame
TyPlDf: TypeAlias = pl.DataFrame

TyArr = list[Any]
TyDic = dict[Any, Any]
TyAoA = list[TyArr]
TyAoD = list[TyDic]
TyDoAoA = dict[Any, TyAoA]
TyDoAoD = dict[Any, TyAoD]
TyDoPdDf = dict[str, TyPdDf] | dict[Any, TyPdDf]
TyDoPlDf = dict[str, TyPlDf] | dict[Any, TyPlDf]
TyDoWsOp = dict[Any, TyWsOp]
TyAoD_DoAoD = TyAoD | TyDoAoD
TyPdDf_DoPdDf = TyPdDf | TyDoPdDf
TyPlDf_DoPlDf = TyPlDf | TyDoPlDf
TyPath = str
TyPathnm = str

TyPlSheetsId = int | Sequence[int] | Literal[0]
TyPlSheetsNm = str | list[str] | tuple[str]
TyPlSheets = TyPlSheetsId | TyPlSheetsNm

TySheet = int | str
TySheets = int | str | list[int | str]
TySheetname = str
TySheetnames = list[TySheetname]

TnArr = None | TyArr
TnAoA = None | TyAoA
TnAoD = None | TyAoD
TnAoD_DoAoD = None | TyAoD_DoAoD
TnDic = None | TyDic
TnDoAoA = None | TyDoAoA
TnDoAoD = None | TyDoAoD
TnDoWsOp = None | TyDoWsOp
TnPdDf = None | TyPdDf
TnPdDf_DoPdDf = None | TyPdDf_DoPdDf
TnDoPdDf = None | TyDoPdDf
TnDoPlDf = None | TyDoPlDf
TnPlDf = None | TyPlDf
TnPlDf_DoPlDf = None | TyPlDf_DoPlDf
TnPlSheetsId = None | TyPlSheetsId
TnPlSheetsNm = None | TyPlSheetsNm
TnPlSheets = None | TyPlSheets
TnSheet = None | TySheet
TnSheets = None | TySheets
TnSheetname = None | TySheetname
TnSheetnames = None | TySheetnames
TnWsOp = None | TyWsOp
TnDf_DoDf = TnPdDf_DoPdDf | TnPlDf_DoPlDf


class IoiPathWbPd:

    @staticmethod
    def verify_obj(obj: TnDf_DoDf, path: TyPath, sheet: TnSheets) -> None:
        if obj is not None:
            return
        if sheet is None:
            msg = f"Excel Workbook {path} contains no sheets"
        else:
            msg = f"Sheets {sheet} are not contained in Excel Workbook {path}"
        raise Exception(msg)

    @classmethod
    def read_wb_to_aod(
            cls, path: TyPath, sheet: TnSheet, **kwargs) -> TnAoD:
        return PdDf.to_aod(cls.read_wb_to_df(path, sheet, **kwargs))

    @classmethod
    def read_wb_to_doaod(cls, path: TyPath, sheet: TnSheet, **kwargs) -> TnDoAoD:
        obj: TnDoPdDf = cls.read_wb_to_dodf(path, sheet, **kwargs)
        if not isinstance(obj, dict):
            raise Exception(f"Object: {obj} should be of type dict")
        return DoPdDf.to_doaod(obj)

    @classmethod
    def read_wb_to_aod_or_doaod(
            cls, path: TyPath, sheet: TnSheet, **kwargs) -> TnAoD_DoAoD:
        obj: TnPdDf_DoPdDf = cls.read_wb_to_df_or_dodf(path, sheet, **kwargs)
        if isinstance(obj, dict):
            return DoPdDf.to_doaod(obj)
        return PdDf.to_aod(obj)

    @classmethod
    def read_wb_to_df(cls, path: TyPath, sheet: TnSheet, **kwargs) -> TnPdDf:
        Path.verify(path)
        if not isinstance(sheet, (int, str)):
            msg = f"sheet; {sheet}  must be of type (int, str)"
            raise Exception(msg)
        _obj: TnPdDf = pd.read_excel(path, sheet_name=sheet, **kwargs)
        cls.verify_obj(_obj, path, sheet)
        return _obj

    @classmethod
    def read_wb_to_df_or_dodf(
            cls, path: TyPath, sheet: TnSheet, **kwargs) -> TnPdDf_DoPdDf:
        Path.verify(path)
        if not (sheet is None or isinstance(sheet, (int, str, list, tuple))):
            msg = f"sheet; {sheet} must be None or of type (int, str, list, tuple)"
            raise Exception(msg)
        obj: TnPdDf_DoPdDf = pd.read_excel(path, sheet_name=sheet)
        cls.verify_obj(obj, path, sheet)
        return obj

    @classmethod
    def read_wb_to_dodf(cls, path: TyPath, sheet: TnSheet, **kwargs) -> TnDoPdDf:
        Path.verify(path)
        if not (sheet is None or isinstance(sheet, (list, tuple))):
            msg = f"sheet; {sheet} must be None or of type (list, tuple)"
            raise Exception(msg)
        _obj: TnPdDf_DoPdDf = pd.read_excel(path, sheet_name=sheet, **kwargs)
        if not isinstance(_obj, dict):
            raise Exception(f"Object: {_obj} should be of type dict")
        return _obj


class IoiPathWbPl:

    @staticmethod
    def is_str_list(val: Iterator) -> TypeGuard[list[str]]:
        '''Determines whether all objects in the list are strings'''
        return all(isinstance(x, str) for x in val)

    @staticmethod
    def is_int_list(val: Iterator) -> TypeGuard[list[int]]:
        '''Determines whether all objects in the list are integers'''
        return all(isinstance(x, int) for x in val)

    @classmethod
    def sh_obj_for_iterator(
            cls, path: TyPath, sheet: Iterator, **kwargs) -> TnPlDf_DoPlDf:
        if cls.is_int_list(sheet):
            return pl.read_excel(path, sheet_id=sheet, **kwargs)
        elif cls.is_str_list(sheet):
            return pl.read_excel(path, sheet_id=None, sheet_name=sheet, **kwargs)
        else:
            raise Exception(f"sheet; {sheet} is is not of Type list[int] or list[str]")

    @staticmethod
    def verify_obj(obj: TnDf_DoDf, path: TyPath, sheet: TnPlSheets) -> None:
        if obj is not None:
            return
        if sheet is None:
            msg = f"Excel Workbook {path} contains no sheets"
        else:
            msg = f"Sheets {sheet} are not contained in Excel Workbook {path}"
        raise Exception(msg)

    @classmethod
    def read_wb_to_aod(cls, path: TyPath, sheet: TnPlSheets, **kwargs) -> TnAoD:
        _obj: TnPlDf = cls.read_wb_to_df(path, sheet, **kwargs)
        return PlDf.to_aod(_obj)

    @classmethod
    def read_wb_to_aod_or_doaod(
            cls, path: TyPath, sheet: TnPlSheets, **kwargs) -> TnAoD_DoAoD:
        _obj: TnPlDf_DoPlDf = cls.read_wb_to_df_or_dodf(path, sheet, **kwargs)
        if isinstance(_obj, dict):
            return DoPlDf.to_doaod(_obj)
        return PlDf.to_aod(_obj)

    @classmethod
    def read_wb_to_doaod(cls, path: TyPath, sheet: TnPlSheets, **kwargs) -> TnDoAoD:
        _obj: TnDoPlDf = cls.read_wb_to_dodf(path, sheet, **kwargs)
        return DoPlDf.to_doaod(_obj)

    @classmethod
    def read_wb_to_df(cls, path: TyPath, sheet: TnPlSheets, **kwargs) -> TnPlDf:
        Path.verify(path)
        if isinstance(sheet, str):
            _obj: TnPlDf = pl.read_excel(
                    path, sheet_id=None, sheet_name=sheet, **kwargs)
        elif isinstance(sheet, int):
            if sheet == 0:
                raise Exception(f"sheet; {sheet} should not be 0")
            _obj = pl.read_excel(path, sheet_id=sheet, **kwargs)
        else:
            raise Exception(f"sheet; {sheet} is invalid")
        cls.verify_obj(_obj, path, sheet)
        return _obj

    @classmethod
    def read_wb_to_df_or_dodf(
            cls, path: TyPath, sheet: TnPlSheets, **kwargs) -> TnPlDf_DoPlDf:
        Path.verify(path)
        if isinstance(sheet, str):
            _obj: TnPlDf_DoPlDf = pl.read_excel(
                    path, sheet_id=None, sheet_name=sheet, **kwargs)
        elif isinstance(sheet, int):
            _obj = pl.read_excel(path, sheet_id=sheet, **kwargs)
        elif isinstance(sheet, Iterator):
            _obj = cls.sh_obj_for_iterator(path, sheet, **kwargs)
        else:
            raise Exception(f"sheet; {sheet} is invalid")
        cls.verify_obj(_obj, path, sheet)
        return _obj

    @classmethod
    def read_wb_to_dodf(cls, path: TyPath, sheet: TnPlSheets, **kwargs) -> TnDoPlDf:
        Path.verify(path)
        if isinstance(sheet, str):
            _obj: TnDoPlDf = pl.read_excel(
                    path, sheet_id=None, sheet_name=[sheet], **kwargs)
        elif isinstance(sheet, int):
            if sheet == 0:
                _obj = pl.read_excel(path, sheet_id=0, **kwargs)
            else:
                _obj = pl.read_excel(path, sheet_id=[sheet], **kwargs)
        elif isinstance(sheet, Iterator):
            _obj = cls.sh_obj_for_iterator(path, sheet, **kwargs)
        else:
            raise Exception(f"sheet; {sheet} is invalid")
        cls.verify_obj(_obj, path, sheet)
        if not isinstance(_obj, dict):
            raise Exception(f"Object: {_obj} should be of type dict")
        return _obj


class IoiPathWbOp:

    @staticmethod
    def load(path: str, **kwargs) -> TyWbOp:
        if path == '':
            raise Exception('path is empty String')
        if path is None:
            raise Exception('path is None')
        try:
            wb: TyWbOp = op.load_workbook(path, **kwargs)
        except Exception as e:
            msg = f"openpyxl.load_workbook for path = {path} throw exception {e}"
            raise Exception(msg)
        return wb

    @classmethod
    def read_wb_to_aod(cls, path: TyPath, sheet: TnSheet, **kwargs) -> TyAoD:
        Path.verify(path)
        _wb: TyWbOp = cls.load(path, **kwargs)
        return WbOp.to_aod(_wb, sheet)

    @staticmethod
    def read_wb_to_doaod(
            path: TyPath, sheet: TnSheets, **kwargs) -> TyDoAoD:
        Path.verify(path)
        _wb: TyWbOp = IoiPathWbOp.load(path, **kwargs)
        return WbOp.to_doaod(_wb, sheet)

    @staticmethod
    def read_wb_to_aod_or_doaod(
            path: TyPath, sheet: TnSheets, **kwargs) -> TnAoD_DoAoD:
        Path.verify(path)
        _wb: TyWbOp = IoiPathWbOp.load(path, **kwargs)
        return WbOp.to_aod_or_doaod(_wb, sheet)

    @staticmethod
    def read_wb_to_aoa(path: TyPath, **kwargs) -> tuple[TyAoA, TyAoA]:
        Path.verify(path)
        wb: TyWbOp = IoiPathWbOp.load(path)
        heads_sheet_name = kwargs.get('headers_sheet_name')
        ws_names: TySheetnames = WbOp.sh_sheetnames(wb, **kwargs)
        aoa = []
        if heads_sheet_name is not None:
            ws = wb[heads_sheet_name]
            heads = WsOp.sh_headers(ws, **kwargs)
        else:
            heads = []
        for ws_name in ws_names:
            Log.Eq.debug("ws_name", ws_name)
            ws = wb[ws_name]
            aoa_ws = WsOp.sh_aoa(ws, sheet_name=ws_name, **kwargs)
            aoa.extend(aoa_ws)
            Log.Eq.debug("aoa_ws", aoa_ws)
        return heads, aoa

    @classmethod
    def read_wb_to_aoa_using_prefix(cls, **kwargs) -> TyAoA:
        # ex_read_workbook_2_aoa(cls, **kwargs):
        # def ex_read_aoa(cls, **kwargs):
        prefix = kwargs.get('prefix')
        if prefix is not None:
            prefix = f"_{prefix}"
        in_path: TyPath = kwargs.get(f'in_path{prefix}', '')
        row_start = kwargs.get(f'row_start{prefix}')
        cols_count = kwargs.get(f'cols_count{prefix}')
        sw_add_sheet_name = kwargs.get(f'sw_add_sheet_name{prefix}')
        sheet_names = kwargs.get(f'sheet_names{prefix}')
        headers_sheet_name = kwargs.get(f'headers_sheet_name{prefix}')
        headers_start = kwargs.get(f'headers_start{prefix}')
        Path.verify(in_path)
        heads, aoa = cls.read_wb_to_aoa(
                in_path,
                row_start=row_start,
                cols_count=cols_count,
                sw_add_sheet_name=sw_add_sheet_name,
                sheet_names=sheet_names,
                headers_sheet_name=headers_sheet_name,
                headers_start=headers_start)
        return aoa


class IoiPathWsOp:

    @staticmethod
    def read_ws_to_dic(
            path: str, sheet: TySheet) -> TnDic:
        _wb: TyWbOp = IoiPathWbOp.load(path)
        _ws: TnWsOp = WbOp.sh_sheet(_wb, sheet)
        return WsOp.to_dic(_ws)

    @staticmethod
    def read_ws_to_aod(
            path: str, sheet: TySheet) -> TnAoD:
        _wb: TyWbOp = IoiPathWbOp.load(path)
        _ws: TnWsOp = WbOp.sh_sheet(_wb, sheet)
        return WsOp.to_aod(_ws)

    @staticmethod
    def read_ws_filter_rows(path, sheet: TySheet) -> TnArr:
        Path.verify(path)
        _wb: TyWbOp = IoiPathWbOp.load(path)
        _ws: TnWsOp = WbOp.sh_sheet(_wb, sheet)
        return WsOp.filter_rows(_ws)

    @staticmethod
    def read_ws_to_aoa(
            path: str, sheet: TnSheets = None) -> tuple[TnAoA, TnSheetnames]:
        Path.verify(path)
        _wb: TyWbOp = IoiPathWbOp.load(path)
        aoa: TyAoA = []
        if not sheet:
            return aoa, None
        _sheetnames: TnSheetnames = WbOp.sh_sheetnames(_wb, sheet)
        if not _sheetnames:
            return aoa, _sheetnames
        for _sheetname in _sheetnames:
            _ws: TnWsOp = WbOp.sh_worksheet(_wb, _sheetname)
            if _ws is not None:
                values: TyArr = WsOp.to_row_values(_ws)
                aoa.append(values)
        return aoa, _sheetnames

    @staticmethod
    def read_sheetnames(path: str) -> TyArr:
        Path.verify(path)
        wb: TyWbOp = IoiPathWbOp.load(path)
        sheetnames: TySheetnames = wb.sheetnames
        return sheetnames

    @staticmethod
    def read_ws_to_doaoa(
            path: str, sheet: TnSheets = None) -> tuple[TnDoAoA, TnSheetnames]:
        Path.verify(path)
        _wb: TyWbOp = IoiPathWbOp.load(path)
        doaoa: TyDoAoA = {}
        if _wb is None:
            return doaoa, None
        sheetnames: TnSheetnames = WbOp.sh_sheetnames(_wb, sheet)
        if not sheetnames:
            return doaoa, sheetnames
        for _sheetname in sheetnames:
            _ws: TnWsOp = WbOp.sh_worksheet(_wb, _sheetname)
            if _ws is not None:
                values: TyArr = WsOp.to_row_values(_ws)
                doaoa[sheet] = values
        return doaoa, sheetnames

    @staticmethod
    def read_ws_to_dowsop(
            path: str, sheet: TnSheets = None) -> tuple[TnDoWsOp, TnSheetnames]:
        Path.verify(path)
        _wb: TyWbOp = IoiPathWbOp.load(path)
        dows: TyDoWsOp = {}
        if _wb is None:
            return dows, None
        sheetnames: TnSheetnames = WbOp.sh_sheetnames(_wb, sheet)
        if not sheetnames:
            return dows, sheetnames
        for _sheetname in sheetnames:
            _ws: TnWsOp = WbOp.sh_worksheet(_wb, _sheetname)
            if _ws is not None:
                dows[_sheetname] = _ws
        return dows, sheetnames
