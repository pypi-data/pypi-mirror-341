
# This file is part of the pdfsp project
# Copyright (C) 2025 Sermet Pekin
#
# This source code is free software; you can redistribute it and/or
# modify it under the terms of the European Union Public License
# (EUPL), Version 1.2, as published by the European Commission.
#
# You should have received a copy of the EUPL version 1.2 along with this
# program. If not, you can obtain it at:
# <https://joinup.ec.europa.eu/collection/eupl/eupl-text-eupl-12>.
#
# This source code is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# European Union Public License for more details.
#
# Alternatively, if agreed upon, you may use this code under any later
# version of the EUPL published by the European Commission.
import pdfplumber
import pandas as pd
from dataclasses import dataclass
import traceback
import os
from pathlib import Path
@dataclass
class DataFrame:
    """A class to handle DataFrame operations."""
    df: pd.DataFrame
    path: Path
    out: str = None
    index: int = 1
    extra: tuple = ()
    name: str = ""
    def __post_init__(self):
        if self.out is None:
            self.out = "Output"
        self.out = Path(self.out)
        self.df = self.make_unique_cols(self.df)
        self.name = Path(self.path).stem.split(".pdf")[0]
    def make_unique_cols(self, df: pd.DataFrame) -> pd.DataFrame:
        cols = [str(x) for x in df.columns]
        df.columns = self.make_unique(cols)
        return df
    def make_unique(self, cols: list[str]) -> list[str]:
        a = []
        b = []
        for i, col in enumerate(cols):
            ucol = col
            if col in a:
                col = col + str(i)
                ucol = f"{col}-{i}"
            a.append(col)
            b.append(ucol)
        return b
    def get_file_name(self):
        file_name = f"[{self.name}]-{self.index}.xlsx"
        return file_name
    def write2(self) -> None:
        self.create_dir()
        file_name = self.get_file_name()
        try:
            self.df.to_excel(self.out / file_name, index=False)
            print(f"[writing table] {file_name}")
        except Exception:
            traceback.print_exc()
    def create_dir(self):
        os.makedirs(self.out, exist_ok=True)
    def write(self):
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        self.create_dir()
        file_name = self.get_file_name()
        wb = Workbook()
        ws = wb.active
        title = f"{self.name}-Table-{self.index}"
        ws.title = title
        # Write the DataFrame to the worksheet
        for r_idx, row in enumerate(
            dataframe_to_rows(self.df, index=False, header=True), start=1
        ):
            ws.append(row)
        # Add a footnote below the table
        footnote_row = len(self.df) + 4  # Position below the DataFrame
        ws.cell(row=footnote_row, column=1, value=f"Footnote: {title} ")
        # Add a paragraph (multi-line text)
        paragraph_row = footnote_row + 4  # Position below the footnote
        ws.cell(
            row=paragraph_row,
            column=1,
            value=f"This table was extracted from {self.path} with pdfsp package.",
        )
        # Save the workbook
        wb.save(self.out / file_name)
        print(f"[writing table] {file_name}")
def get_pdf_files(folder: Path = None, out: str = None) -> list[str]:
    """Get all PDF files in the specified folder."""
    if folder is None:
        folder = Path(".")
    if out is None:
        out = Path(".")
    files = [x for x in os.listdir() if x.endswith(".pdf")]
    if not files:
        print(f"No PDF files found in {folder}")
        return []
    return files
def extract_tables_from_pdf(pdf_path, out: Path = None) -> list[DataFrame]:
    """Extract tables from a PDF file."""
    pdfs = []
    with pdfplumber.open(pdf_path) as pdf:
        print(f"""Extracting tables from {pdf_path}""")
        for _, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            for index, table in enumerate(tables):
                df = pd.DataFrame(table[1:], columns=table[0])
                pdfs.append(DataFrame(df, pdf_path, out, index=index + 1))
    return pdfs
def write_dfs(pdf_files: list[Path], out: Path = None):
    """Write DataFrames to Excel files."""
    for pdf_file in pdf_files:
        pdfs: list[DataFrame] = extract_tables_from_pdf(pdf_file, out)
        for df in pdfs:
            df.write()
def extract_tables(folder: Path = None, out: str = None):
    """Extract tables from all PDF files in the specified folder."""
    files = get_pdf_files(folder, out)
    write_dfs(files, out)
# ........................................ main
# ........................................
# pdf_paths = ["aa.pdf", "bb.pdf"]
# write_dfs(pdf_paths)
# ........................................
# write_tables_folder(".", "a3")