#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Converter module for transforming test cases to CSV and Excel formats.
"""

import os
import csv
from typing import Dict, List, Any, Optional
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from loguru import logger


class TestCaseConverter:
    """Converter for transforming test cases to various formats."""

    def __init__(self, output_dir: str = "output"):
        """
        Initialize the converter.

        Args:
            output_dir: Directory to store output files.
        """
        self.output_dir = output_dir
        self._ensure_output_dir()
        
    def _ensure_output_dir(self):
        """Ensure the output directory exists."""
        os.makedirs(self.output_dir, exist_ok=True)
        
    def convert_to_csv(self, test_cases: Dict[str, List[Dict[str, Any]]], force: bool = False) -> List[str]:
        """
        Convert test cases to CSV files.
        
        Args:
            test_cases: Dictionary with test case file names as keys and lists of test case dictionaries as values.
            force: Whether to overwrite existing files without asking.
            
        Returns:
            List of paths to the created CSV files.
        """
        csv_files = []
        
        for file_name, cases in test_cases.items():
            if not cases:
                logger.warning(f"No test cases to convert for {file_name}")
                continue
                
            csv_path = os.path.join(self.output_dir, file_name)
            
            # Check if file exists and prompt for overwrite if not forced
            if os.path.exists(csv_path) and not force:
                logger.warning(f"File {csv_path} already exists. Use --force to overwrite.")
                continue
                
            # Get field names from the first test case
            fieldnames = list(cases[0].keys())
            
            try:
                with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(cases)
                
                logger.info(f"Created CSV file: {csv_path}")
                csv_files.append(csv_path)
                
            except Exception as e:
                logger.error(f"Error creating CSV file {csv_path}: {str(e)}")
        
        return csv_files
    
    def convert_to_excel(self, test_cases: Dict[str, List[Dict[str, Any]]], force: bool = False) -> Optional[str]:
        """
        Convert test cases to an Excel file with multiple sheets.
        
        Args:
            test_cases: Dictionary with test case file names as keys and lists of test case dictionaries as values.
            force: Whether to overwrite existing files without asking.
            
        Returns:
            Path to the created Excel file, or None if no file was created.
        """
        if not test_cases:
            logger.warning("No test cases to convert to Excel")
            return None
            
        excel_path = os.path.join(self.output_dir, "test_cases.xlsx")
        
        # Check if file exists and prompt for overwrite if not forced
        if os.path.exists(excel_path) and not force:
            logger.warning(f"File {excel_path} already exists. Use --force to overwrite.")
            return None
            
        try:
            wb = openpyxl.Workbook()
            
            # Remove the default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            for file_name, cases in test_cases.items():
                if not cases:
                    logger.warning(f"No test cases to add to Excel for {file_name}")
                    continue
                    
                # Create a new sheet for this file
                sheet_name = Path(file_name).stem
                sheet = wb.create_sheet(title=sheet_name)
                
                # Get field names from the first test case
                fieldnames = list(cases[0].keys())
                
                # Add header row
                for col_idx, field in enumerate(fieldnames, 1):
                    cell = sheet.cell(row=1, column=col_idx, value=field)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Add data rows
                for row_idx, case in enumerate(cases, 2):
                    for col_idx, field in enumerate(fieldnames, 1):
                        value = case.get(field, "")
                        sheet.cell(row=row_idx, column=col_idx, value=value)
                
                # Auto-adjust column widths
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter  # Get column letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    sheet.column_dimensions[column].width = adjusted_width
            
            wb.save(excel_path)
            logger.info(f"Created Excel file: {excel_path}")
            return excel_path
            
        except Exception as e:
            logger.error(f"Error creating Excel file {excel_path}: {str(e)}")
            return None
