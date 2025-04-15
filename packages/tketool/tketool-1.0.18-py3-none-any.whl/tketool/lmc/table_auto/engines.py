import abc
import pandas as pd
import os.path
import openpyxl
import re, math, json
from tketool.lmc.prompts.prompt_controller import get_prompt, get_prompt_by_path
from tketool.utils.progressbar import process_status_bar
from tketool.utils.MultiTask import do_multitask
from tketool.lmc.lmc_linked import *
from tketool.lmc.lmc_linked_flow import lmc_linked_flow_model
from openpyxl import load_workbook
from openpyxl.styles import Alignment


class excel_pointer:
    """
    Excel指针类，用于定位Excel中的特定单元格。

    参数:
    - sheet_name: 工作表名称
    - col_name: 列名
    - col_excel_name: Excel中的列名（如A, B, C）
    - row_index: 行索引
    """

    def __init__(self, sheet_name, col_name, col_excel_name, row_index):
        self.sheet_name = sheet_name
        self.col_name = col_name
        self.col_excel_name = col_excel_name
        self.row_index = row_index


# Excel代理基类,定义了Excel处理的基本接口
# 使用abc.ABCMeta作为元类来定义抽象基类
class excel_agent(metaclass=abc.ABCMeta):
    """
    Excel代理基类，定义了Excel处理的基本接口。

    属性:
    - match_str: 用于匹配的字符串
    - params_list: 参数列表
    - agent_des: 代理描述

    方法:
    - call: 执行代理的主要方法
    - init_task: 初始化任务
    - is_nan: 检查值是否为NaN
    - get_datas_by: 根据索引获取数据
    """

    @property
    @abc.abstractmethod
    def match_str(self):
        pass

    @property
    @abc.abstractmethod
    def params_list(self):
        return

    @property
    def agent_des(self):
        return ""

    @abc.abstractmethod
    def call(self, llm, row_dict, cur_col_name, sheet_obj, params, content, logs_list):
        """
        执行代理的主要方法。

        参数:
        - llm: 语言模型实例
        - row_dict: 当前行的数据字典
        - cur_name: 当前列名
        - sheet_obj: 工作表对象
        - params: 参数
        - content: 内容
        - logs_list: 日志列表
        """
        pass

    def init_task(self, llm, sheet_obj, task_define_loc, params, content):
        """
        初始化任务。

        参数:
        - llm: 语言模型实例
        - sheet_obj: 工作表对象
        - params: 参数
        - content: 内容
        """
        pass

    def is_nan(self, value):
        """
        检查值是否为NaN。

        参数:
        - value: 要检查的值

        返回:
        - bool: 如果是NaN则返回True，否则返回False
        """
        return value is None or value == '' or (isinstance(value, float) and math.isnan(value))

    def get_datas_by(self, sheet_obj, index_key, include_header=False, cur_col=None):
        """
        根据索引获取数据。

        参数:
        - sheet_obj: 工作表对象
        - index_key: 索引键
        - cur_col: 当前列（可选）

        返回:
        - dict或list: 返回行数据或列数据
        """

        def split_string(s):
            # 使用正则表达式匹配字符和数字
            letters = re.findall(r'[A-Za-z]+', s)
            digits = re.findall(r'\d+', s)
            return [letters[0] if letters else '', digits[0] if digits else '']

        sheet_name = cur_col.sheet_name if cur_col is not None else ""

        sheet_split_list = index_key.split("::")
        if len(sheet_split_list) > 1:
            sheet_name = sheet_split_list[0]
            r_c_key = sheet_split_list[1]
        else:
            r_c_key = sheet_split_list[0]

        rc = split_string(r_c_key)

        if rc[0] == '':
            row_index = int(rc[1])
            row_data = {c: v
                        for v, c in
                        zip(sheet_obj[sheet_name].rows[row_index], sheet_obj[sheet_name].excel_column_names)}
            return row_data
        elif rc[1] == '':
            col_index = sheet_obj[sheet_name].excel_column_names.index(rc[0])
            col_list = []
            if include_header:
                col_list.append(sheet_obj[sheet_name].column_names[col_index])
            col_list += [s[col_index] for s in sheet_obj[sheet_name].rows]
            return col_list
        else:
            col_index = sheet_obj[sheet_name].excel_column_names.index(rc[0])
            row_index = int(rc[1])
            return sheet_obj[sheet_name].rows[row_index][col_index]


class double_shape(excel_agent):
    """
    双形代理类，继承自excel_agent。

    属性:
    - params_list: 参数列表
    - match_str: 用于匹配的字符串

    方法:
    - call: 执行代理的主要方法
    """

    @property
    def params_list(self):
        return []

    @property
    def match_str(self):
        return "##"

    def call(self, llm, row_dict, cur_name, sheet_obj, params, content, logs_list):
        """
        执行代理的主要方法。

        参数:
        - llm: 语言模型实例
        - row_dict: 当前行的数据字典
        - cur_name: 当前列名
        - sheet_obj: 工作表对象
        - params: 参数
        - content: 内容
        - logs_list: 日志列表

        返回:
        - str: 结果字符串
        """
        llm_invoker = lmc_linked_model(llm).set_prompt_template(content)
        result = llm_invoker(**row_dict)
        if len(result.results) > 0:
            return str(result.result)
        else:
            return ""

    @property
    def agent_des(self):
        return "基本调用大模型的算子。"


class prompt_file_shape(excel_agent):
    """
    提示文件形状代理类，继承自excel_agent。

    属性:
    - params_list: 参数列表
    - match_str: 用于匹配的字符串

    方法:
    - init_task: 初始化任务
    - call: 执行代理的主要方法
    """

    def __init__(self):
        self.invoker_dict = {}

    @property
    def params_list(self):
        return [("prompt_file_path", "提示词文件路径"),
                ("prompt_key_mapping", "提示词的mapping key替换, A=wordA#B=wordB"),
                ("prompt_model_output", "返回的结果结构选择器(可选)"), ]

    @property
    def match_str(self):
        return "#promptfile"

    def init_task(self, llm, sheet_obj, task_define_loc, params, content):
        """
        初始化任务。

        参数:
        - llm: 语言模型实例
        - sheet_obj: 工作表对象
        - params: 参数
        - content: 内容
        """
        invoker_key = params["prompt_file_path"]
        if invoker_key not in self.invoker_dict:
            path = params["prompt_file_path"]
            parse_field = get_prompt_by_path(path)
            self.invoker_dict[invoker_key] = lmc_linked_flow_model(parse_field, retry_time=2)

    def call(self, llm, row_dict, cur_name, sheet_obj, params, content, logs_list):
        """
        执行代理的主要方法。

        参数:
        - llm: 语言模型实例
        - row_dict: 当前行的数据字典
        - cur_name: 当前列名
        - sheet_obj: 工作表对象
        - params: 参数
        - content: 内容
        - logs_list: 日志列表

        返回:
        - str: 结果字符串
        """
        invoker_key = params["prompt_file_path"]
        mapping_dict = {}
        for k, v in row_dict.items():
            if isinstance(v, str):
                mapping_dict[k] = v
            else:
                if not self.is_nan(v):
                    mapping_dict[k] = v
        if len(params) > 1:
            mapping_change_dict = {}
            mapping_change = params["prompt_key_mapping"].split("#")
            for spp in mapping_change:
                s = spp.split("=")
                if len(s) == 2:
                    mapping_change_dict[s[0]] = s[1]

            for k, v in mapping_change_dict.items():
                if k in mapping_dict:
                    mapping_dict[v] = mapping_dict[k]
                if v in mapping_dict:
                    mapping_dict[k] = mapping_dict[v]

        llmresult = self.invoker_dict[invoker_key](llm, **mapping_dict)
        if not llmresult.passed:
            logs_list.append("invoke llm error")
        if len(params) > 2:
            output_str = params["prompt_model_output"]
            result_str = eval("llmresult.result." + output_str)
            return result_str
        else:
            return llmresult.result.json()

    @property
    def agent_des(self):
        return "用于处理提示文件。"


class sheet_data:
    """
    工作表数据类，用于处理Excel工作表的数据。

    参数:
    - sheet_obj: 工作表对象
    - all_agents: 所有代理的字典

    方法:
    - parse_colstr: 解析列字符串
    - set: 设置单元格的值
    """

    def __init__(self, sheet_obj, all_agents):
        self.sheet_obj = sheet_obj
        self.all_agents = all_agents
        self.column_names = []
        self.column_name_index = {}
        self.excel_column_names = []
        self.column_agent = []
        self.rows = []
        self.log_col = None

        for col_name in sheet_obj.columns.tolist():
            self.column_names.append(col_name)
            excel_col_name = openpyxl.utils.get_column_letter(sheet_obj.columns.get_loc(col_name) + 1)
            self.excel_column_names.append(excel_col_name)
            self.column_agent.append(self.parse_colstr(col_name))
            if col_name == '#log':
                self.log_col = excel_col_name

        self.mapping_agent = []  # all_agents[cmd] if cmd in all_agents else None for cmd, par, con in self.column_agent]

        for cmd, par, con in self.column_agent:
            if cmd in all_agents:
                self.mapping_agent.append(all_agents[cmd])
            else:
                self.mapping_agent.append(None)

        for idx, row in sheet_obj.iterrows():
            cur_row = []
            for c_idx in self.column_names:
                cur_row.append(row[c_idx])
            self.rows.append(cur_row)

        self.column_name_index = {colname: idx for idx, colname in enumerate(self.column_names)}

    def parse_colstr(self, col_str):
        """
        解析列字符串。

        参数:
        - col_str: 列字符串

        返回:
        - tuple: (命令, 参数, 内容)
        """
        if not isinstance(col_str, str):
            return (None, None, None)
        if col_str.startswith("##"):
            return ("##", [], col_str[2:])
        else:
            pattern = r'(#\w+)(?:\(([^)]+)\))?(?::\s*(.*))?'
            match = re.match(pattern, col_str)
            if match:
                command = match.group(1).strip()
                params = match.group(2)
                content = match.group(3)
                # 处理参数
                params_list = [p.strip().strip('"').strip("'") for p in params.split(',')] if params else []

                if command in self.all_agents:
                    zip_params = {k[0]: v for k, v in zip(self.all_agents[command].params_list, params_list)}
                    return (command, zip_params, content)

        return (None, None, None)

    def set(self, row_index, col_name, value):
        """
        设置单元格的值。

        参数:
        - row_index: 行索引
        - col_name: 列名
        - value: 要设置的值
        """
        target_col_name = ""
        if col_name in self.column_name_index:
            target_col_name = col_name
        else:
            target_col_name = self.column_names[self.excel_column_names.index(col_name)]
        
        # 确保列的数据类型为object，以便可以存储字符串
        if self.sheet_obj[target_col_name].dtype != 'object':
            self.sheet_obj[target_col_name] = self.sheet_obj[target_col_name].astype('object')
        
        self.sheet_obj.at[row_index, target_col_name] = value
        col_index = self.column_name_index[target_col_name]
        self.rows[row_index][col_index] = value
        pass


class excel_work_process:
    """
    Excel工作进程类，用于跟踪工作进度。

    属性:
    - max_sheet_count: 最大工作表数
    - done_sheet_count: 已完成的工作表数
    """

    def __init__(self):
        self.max_sheet_count = 0
        self.done_sheet_count = 0


class excel_engine:
    """
    Excel引擎类，用于处理Excel文件。

    参数:
    - llm: 语言模型实例
    - args: 代理实例
    - thread: 线程数

    方法:
    - parse_sheet: 解析工作表
    - call_file: 处理Excel文件
    """

    def __init__(self, llm, *args, thread=1):
        self.all_agents = {}
        self.llm = llm
        self.thread_count = thread
        for arg in args:
            if isinstance(arg, excel_agent):
                self.all_agents[arg.match_str] = arg

    def parse_sheet(self, sheet_obj_dict, sheet_key, pass_row_count, pb, progress_callback=None):
        """
        解析工作表。

        参数:
        - sheet_obj_dict: 工作表对象字典
        - sheet_key: 工作表键
        - pass_row_count: 跳过的行数
        - pb: 进度条对象
        - progress_callback: 进度回调函数
        """

        def do_task(idx):
            if progress_callback:
                progress_callback(idx)

            sheet_obj = sheet_obj_dict[sheet_key]
            logs_list = []
            for op, row_tile, coln, coln2 in zip(sheet_obj.mapping_agent, sheet_obj.column_agent,
                                                 sheet_obj.column_names,
                                                 sheet_obj.excel_column_names):
                if op is None:
                    continue

                row_dict = {k: v for k, v in
                            zip(sheet_obj.excel_column_names, sheet_obj.rows[idx])}

                if not isinstance(row_dict[coln2], str) and math.isnan(row_dict[coln2]):
                    params = row_tile[1]
                    content = row_tile[2]
                    pointer = excel_pointer(sheet_key, coln, coln2, idx)
                    call_result = op.call(self.llm, row_dict, pointer, sheet_obj_dict, params, content, logs_list)
                    sheet_obj_dict[sheet_key].set(idx, coln, call_result)
            if sheet_obj.log_col:
                sheet_obj_dict[sheet_key].set(idx, sheet_obj.log_col, "\n".join(logs_list))
            pass

        rows_index = list(range(pass_row_count, len(sheet_obj_dict[sheet_key].rows)))
        for row, c_row in pb.iter_bar(
                do_multitask(rows_index, do_task, self.thread_count, self.thread_count * 2),
                key="row", max=len(rows_index)):
            pass

    def call_file(self, excel_file_path: str, start_row_index=0, progress_callback=None):
        """
        处理Excel文件。

        参数:
        - excel_file_path: Excel文件路径
        - start_row_index: 开始行索引
        - progress_callback: 进度回调函数
        """
        xls = pd.ExcelFile(excel_file_path)
        pb = process_status_bar()

        all_sheets = {}
        column_widths = {}  # 用于存储每个工作表的列宽信息

        total_sheets = len(xls.sheet_names)
        current_sheet = 0

        for sheet in xls.sheet_names:
            sheet_content = sheet_data(xls.parse(sheet), all_agents=self.all_agents)
            all_sheets[sheet] = sheet_content

            # 计算总行数用于进度计算
            total_rows = len(sheet_content.rows) - start_row_index

            # 加载现有的Excel文件，以获取列宽信息
            wb = load_workbook(excel_file_path)
            ws = wb[sheet]

            column_widths[sheet] = {}

            # 获取工作表的列维度
            for col in ws.column_dimensions:
                # 在openpyxl中，列宽默认是None，如果未设置
                # 所以我们需要检查是否已经设置   列宽
                if ws.column_dimensions[col].width is not None:
                    column_widths[sheet][col] = ws.column_dimensions[col].width
                else:
                    # 如果列宽未设置，可以设置一个默认值或根据内容计算宽度
                    column_widths[sheet][col] = 10  # 假设默认列宽为10

            if progress_callback:
                progress_callback({
                    'stage': 'init',
                    'current_sheet': sheet,
                    'sheet_progress': current_sheet / total_sheets,
                    'total_sheets': total_sheets,
                    'current_sheet_index': current_sheet,
                    'total_rows': total_rows
                })
            current_sheet += 1

        for k, sheet in all_sheets.items():
            for cmd, par, con in sheet.column_agent:
                if cmd in self.all_agents:
                    oper_loc = excel_pointer(k, None, None, -1)
                    self.all_agents[cmd].init_task(self.llm, all_sheets, oper_loc, par, con)

        current_sheet = 0
        for s_name, op_sheet in pb.iter_bar(all_sheets.items(), key="sheet"):
            try:
                total_rows = len(op_sheet.rows) - start_row_index
                current_row = 0

                def row_progress_callback(row_index):
                    nonlocal current_row
                    current_row = row_index - start_row_index
                    if progress_callback:
                        progress_callback({
                            'stage': 'processing',
                            'current_sheet': s_name,
                            'sheet_progress': current_sheet / total_sheets,
                            'row_progress': current_row / total_rows,
                            'current_row': current_row,
                            'total_rows': total_rows,
                            'total_sheets': total_sheets,
                            'current_sheet_index': current_sheet
                        })

                self.parse_sheet(all_sheets, s_name, start_row_index, pb=pb, progress_callback=row_progress_callback)
            finally:
                current_sheet += 1

        with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
            for s_name, op_sheet in pb.iter_bar(all_sheets.items(), key="sheet"):
                all_sheets[s_name].sheet_obj.to_excel(writer, sheet_name=s_name, index=False)
                # 设置列宽和自动换行
                worksheet = writer.sheets[s_name]
                for col, width in column_widths[s_name].items():
                    worksheet.column_dimensions[col].width = width
                for col in worksheet.columns:
                    for cell in col:
                        cell.alignment = Alignment(wrapText=True)

    def get_agent_doc(self):
        """
        获取所有代理的文档信息。

        返回:
        - dict: 包含代理名称、介绍和参数信息的字典
        """
        agent_docs = {}
        for match_str, agent in self.all_agents.items():
            agent_docs[match_str] = {
                "name": agent.__class__.__name__,
                "description": agent.agent_des,
                "params": {param[0]: param[1] for param in agent.params_list}
            }
        return agent_docs


class summary_shape(excel_agent):
    """
    文本摘要算子类，继承自excel_agent。
    用于生成文本内容的摘要。

    属性:
    - params_list: 参数列表，包含字数限制参数和输入列名
    - match_str: 用于匹配的字符串 "#summary"
    """

    def __init__(self):
        self.invoker = None

    @property
    def params_list(self):
        return [("input_column", "输入列名"), ("word_limit", "摘要字数限制"), ]

    @property
    def match_str(self):
        return "#summary"

    def init_task(self, llm, sheet_obj, task_define_loc, params, content):
        """初始化任务，创建LMC模型实例"""
        if self.invoker is None:
            parse_field = get_prompt("text_summary", lang="chinese")
            self.invoker = lmc_linked_flow_model(parse_field, retry_time=2)

    def call(self, llm, row_dict, cur_name, sheet_obj, params, content, logs_list):
        """
        执行摘要生成。

        参数:
        - llm: 语言模型实例
        - row_dict: 当前行的数据字典
        - cur_name: 当前列名
        - sheet_obj: 工作表对象
        - params: 参数字典，包含word_limit和input_column
        - content: 要进行摘要的内容
        - logs_list: 日志列表

        返回:
        - str: 生成的摘要文本
        """
        input_column = params.get("input_column")
        if not input_column:
            logs_list.append("未指定输入列")
            return ""

        # 使用get_datas_by方法获取输入列的数据
        input_data = row_dict[input_column]

        # 准备参数
        input_params = {
            "content": input_data,
            "word_limit": params.get("word_limit", "200")  # 默认200字
        }

        # 调用LLM生成摘要
        llm_result = self.invoker(llm, **input_params)

        if not llm_result.passed:
            logs_list.append("生成摘要失败")
            return ""

        return llm_result.result

    @property
    def agent_des(self):
        return "文本摘要算子类，用于生成文本内容的摘要。"


class classification_shape(excel_agent):
    """
    分类算子类，继承自excel_agent。
    用于对文本内容进行分类。

    属性:
    - params_list: 参数列表，包含类别、类别介绍和输入列名
    - match_str: 用于匹配的字符串 "#classification"
    """

    def __init__(self):
        self.invoker = None
        self.category_mapping = {}

    @property
    def params_list(self):
        return [("input_column", "输入列名"),
                ("category", "分类的类别"),
                ("category_description", "类别的介绍（可选）"), ]

    @property
    def match_str(self):
        return "#classification"

    def init_task(self, llm, sheet_obj, task_define_loc, params, content):
        """初始化任务，创建LMC模型实例"""
        if self.invoker is None:
            parse_field = get_prompt("classification", lang="chinese")
            self.invoker = lmc_linked_flow_model(parse_field, retry_time=2)

        # 初始化类别映射
        category = params.get("category", "")
        category_description_column = params.get("category_description", "")

        # 获取输入列和类别描述的数据
        input_data = self.get_datas_by(sheet_obj, category, cur_col=task_define_loc, include_header=True)
        if category_description_column:
            category_description_data = self.get_datas_by(sheet_obj, category_description_column, include_header=True)
        else:
            category_description_data = ["" for _ in input_data]

        # 将结果保存到category_mapping中
        self.category_mapping = {}
        for i in range(len(input_data)):
            if not self.is_nan(input_data[i]):
                self.category_mapping[1000 + i] = (input_data[i], category_description_data[i])
        self.category_content = "\n".join([f"{k}: {v1}, {v2}" for k, (v1, v2) in self.category_mapping.items()])
        pass

    def call(self, llm, row_dict, cur_name, sheet_obj, params, content, logs_list):
        """
        执行分类操作。

        参数:
        - llm: 语言模型实例
        - row_dict: 当前行的数据字典
        - cur_name: 当前列名
        - sheet_obj: 工作表对象
        - params: 参数字典，包含category、category_description和input_column
        - content: 要进行分类的内容
        - logs_list: 日志列表

        返回:
        - str: 分类结果
        """
        input_column = params.get("input_column")
        if not input_column:
            logs_list.append("未指定输入列")
            return ""

        # 使用get_datas_by方法获取输入列的数据
        input_data = row_dict.get(input_column, "")  # self.get_datas_by(sheet_obj, input_column, cur_col=cur_name)

        # 准备参数
        input_params = {
            "content": input_data,
            "category_content": self.category_content,
        }

        # 调用LLM进行分类
        llm_result = self.invoker(llm, **input_params)

        if not llm_result.passed:
            logs_list.append("分类失败")
            return ""

        # 获取分类索引并映射回类别
        classification_index = llm_result.result.classification_index
        result = self.category_mapping.get(classification_index, ("未知类别",""))
        return str(result[0])

    @property
    def agent_des(self):
        return "分类算子类，用于对文本内容进行分类。"
