import os
import datetime
import zoneinfo
from icecream import ic
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Alignment, Protection
from openpyxl.comments import Comment
import uuid
import re


from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError


# If modifying these scopes, delete the file token.json.
# SCOPES = ["https://www.googleapis.com/auth/calendar.readonly"]
SCOPES = ["https://www.googleapis.com/auth/calendar"]

load_dotenv()
CAL_ID = os.getenv("CAL_ID")
if CAL_ID == None:
    CAL_ID = "ja.japanese#holiday@group.v.calendar.google.com"  # 日本の祝日カレンダー



def sign_in():
    """Googleにサインイン"""
    ic(CAL_ID)
    credentials = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists("token.json"):
        credentials = Credentials.from_authorized_user_file("token.json", SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not credentials or not credentials.valid:
        if credentials and credentials.expired and credentials.refresh_token:
            try:
                credentials.refresh(Request())
            except Exception as e:
                print(
                    "ERROR: おそらくtokenの有効期限が切れています。-reオプションを指定して再度実行してください。"
                )
                ic(str(e))
                exit(1)
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", SCOPES)
            credentials = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open("token.json", "w") as token:
            token.write(credentials.to_json())
    return credentials


def get_service():
    """Googleサービスを取得する"""
    credentials = sign_in()
    service = build("calendar", "v3", credentials=credentials)
    return service


def iso2jst(date_str):
    """iso format 文字列から日本時間のdatetimeオブジェクトを作成する"""
    dt = datetime.datetime.fromisoformat(date_str)
    return dt.astimezone(zoneinfo.ZoneInfo("Asia/Tokyo"))


def remove_time_stamp(description):
    """descriptionに記載されている(このシステムで追加した)タイムスタンプを除去する"""
    if description:
        description = re.sub("/////.*時点 /////$", "", description)
        return description.strip()
    else:
        return ""


def append_time_stamp(description):
    """カレンダーの更新日を利用者にわかりやすくするために末尾にタイムスタンプを付与する"""
    today = datetime.datetime.now().strftime("%Y/%m/%d %H:%M")
    description += f"\n\n///// {today} 時点 /////"
    return description


def update_event(service, event):
    """カレンダーにイベントを登録する"""
    result = None
    g_event = None

    if event.get("summary") == "":
        # 削除
        try:
            service.events().delete(
                calendarId=CAL_ID, eventId=f"{event['id']}"
            ).execute()
            return event
        except Exception as e:
            print(f"イベント削除処理でエラーとなりました。{event['id']}")
            ic(e)
    else:
        # 更新
        try:
            # Googleから取得
            g_event = (
                service.events()
                .get(calendarId=CAL_ID, eventId=f"{event['id']}")
                .execute()
            )

            if g_event != None:
                # 値準備
                g_start = g_event["start"].get("dateTime", g_event["start"].get("date"))
                g_end = g_event["end"].get("dateTime", g_event["end"].get("date"))
                g_summary = g_event.get("summary")
                g_description = remove_time_stamp(g_event.get("description"))

                start = event["start"].get("dateTime", event["start"].get("date"))
                end = event["end"].get("dateTime", event["end"].get("date"))
                summary = event.get("summary")
                description = event.get("description")

                # 変更確認
                if (
                    g_start == start
                    and g_end == end
                    and g_summary == summary
                    and g_description == description
                    and g_event["status"] == "confirmed"
                ):
                    # 変更なし
                    pass
                else:
                    # 変更あり。
                    if g_event["status"] != "confirmed":
                        # 削除済みなので更新しない
                        pass
                    else:
                        # 更新する。
                        g_event["start"] = event["start"]
                        g_event["end"] = event["end"]
                        g_event["summary"] = event["summary"]
                        g_event["status"] = "confirmed"
                        g_event["description"] = append_time_stamp(description)
                        result = (
                            service.events()
                            .update(
                                calendarId=CAL_ID, eventId=g_event["id"], body=g_event
                            )
                            .execute()
                        )
                        # debug
                        # ic(g_event, event)
                        if result:
                            print(
                                "Event created/updated: %s" % (result.get("htmlLink"))
                            )
                        else:
                            print("Why result is None?")

        except HttpError as e:
            # 取得できなかったので追加する
            event["description"] = append_time_stamp(event.get("description"))
            result = service.events().insert(calendarId=CAL_ID, body=event).execute()

        return result


def get_events(service, start_date):
    """カレンダーからイベントを取得する"""
    try:
        if start_date:
            ym = start_date.split("-")
            start = datetime.datetime(
                year=int(ym[0]),
                month=int(ym[1]),
                day=1,
                tzinfo=datetime.timezone(datetime.timedelta(hours=9)),
            )
        else:
            now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9)))
            start = datetime.datetime(
                year=now.year,
                month=now.month,
                day=1,
                tzinfo=datetime.timezone(datetime.timedelta(hours=9)),
            )
        print(
            "Getting the upcoming 10 events",
            start.astimezone(datetime.timezone(datetime.timedelta(hours=+9))),
        )
        events_result = (
            service.events()
            .list(
                calendarId=CAL_ID,
                timeMin=start.isoformat(),
                maxResults=10,
                singleEvents=True,
                orderBy="startTime",
            )
            .execute()
        )
        events = events_result.get("items", [])
        return events

    except HttpError as error:
        print(f"An error occurred: {error}")


def download_events(service, school_year, out_file):
    """カレンダーから指定した年度のイベントを取得し、Excelファイルを作成する。"""

    def write_header(ws):
        """シートにヘッダーを記入する"""
        # cell value and comment
        ws.cell(1, 1, "開始日")
        ws.cell(1, 2, "終了日")
        ws.cell(1, 2).comment = Comment(
            text="オプション。記入されていない場合は開始日一日のイベント。指定した場合この日を含むイベントが作成される。",
            author="Kenji Sakurai",
        )
        ws.cell(1, 3, "行事")
        ws.cell(1, 4, "内容")
        ws.cell(1, 5, "作成日")
        ws.cell(1, 5).comment = Comment(
            text="Google側の作成日。更新不可", author="Kenji Sakurai"
        )
        ws.cell(1, 6, "更新日")
        ws.cell(1, 6).comment = Comment(
            text="Google側の更新日。更新不可", author="Kenji Sakurai"
        )
        ws.cell(1, 7, "イベントID")
        ws.cell(1, 7).comment = Comment(
            text="Google側のID。編集不可。", author="Kenji Sakurai"
        )

        # cell protection
        ws.cell(1, 1).protection = Protection(locked=False)
        ws.cell(1, 2).protection = Protection(locked=False)
        ws.cell(1, 3).protection = Protection(locked=False)
        ws.cell(1, 4).protection = Protection(locked=False)
        ws.cell(1, 5).protection = Protection(locked=False)
        ws.cell(1, 6).protection = Protection(locked=False)
        ws.cell(1, 7).protection = Protection(locked=False)

    def write_row(ws, row_num, event):
        """イベントを指定した行に書き出す"""
        id = event["id"]
        st = event["start"].get("dateTime", event["start"].get("date"))
        ed = event["end"].get("dateTime", event["end"].get("date"))
        st = datetime.datetime.fromisoformat(st).replace(tzinfo=None)
        ed = datetime.datetime.fromisoformat(ed).replace(tzinfo=None)
        delta_one = (ed - st) == datetime.timedelta(days=1)
        summary = event["summary"]
        description = event.get("description", "")
        description = remove_time_stamp(description)
        created = iso2jst(event.get("created"))
        updated = iso2jst(event.get("updated"))

        # cell value
        ws.cell(row_num, 1, st)
        if st.strftime("%H:%M:%S") == "00:00:00":
            # date
            ws.cell(row_num, 1).number_format = "yyyy/mm/dd"
        else:
            # datetime
            ws.cell(row_num, 1).number_format = "yyyy/mm/dd hh:mm"
        if delta_one:
            # 1日のイベントの場合、終了日は省略
            pass
        else:
            if ed.strftime("%H:%M:%S") == "00:00:00":
                # 日付のみの終了日を記録する場合、その日を含む、という表現とする
                ed = ed - datetime.timedelta(days=1)
            ws.cell(row_num, 2, ed)
        if ed.strftime("%H:%M:%S") == "00:00:00":
            # date
            ws.cell(row_num, 2).number_format = "yyyy/mm/dd"
        else:
            # datetime
            ws.cell(row_num, 2).number_format = "yyyy/mm/dd hh:mm"

        ws.cell(row_num, 3, summary)
        ws.cell(row_num, 4, description)
        ws.cell(row_num, 5, created.replace(tzinfo=None))
        ws.cell(row_num, 5).number_format = "yyyy/mm/dd hh:mm"
        ws.cell(row_num, 6, updated.replace(tzinfo=None))
        ws.cell(row_num, 6).number_format = "yyyy/mm/dd hh:mm"
        ws.cell(row_num, 7, id)

        # cell alignment
        topleft = Alignment(horizontal="left", vertical="top", wrap_text=False)
        topleft_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.cell(row_num, 1).alignment = topleft
        ws.cell(row_num, 2).alignment = topleft
        ws.cell(row_num, 3).alignment = topleft
        if description.find("\n") >= 0:
            # 折返し
            ws.cell(row_num, 4).alignment = topleft_wrap
        else:
            ws.cell(row_num, 4).alignment = topleft
        ws.cell(row_num, 5).alignment = topleft
        ws.cell(row_num, 6).alignment = topleft
        ws.cell(row_num, 7).alignment = topleft

        # cell protection
        ws.cell(row_num, 1).protection = Protection(locked=False)
        ws.cell(row_num, 2).protection = Protection(locked=False)
        ws.cell(row_num, 3).protection = Protection(locked=False)
        ws.cell(row_num, 4).protection = Protection(locked=False)
        ws.cell(row_num, 5).protection = Protection(locked=True)
        ws.cell(row_num, 6).protection = Protection(locked=True)
        ws.cell(row_num, 7).protection = Protection(locked=True)
        ic(st.isoformat(), summary)

    try:

        wb = openpyxl.Workbook()
        ws = wb.active
        counter = 1

        if school_year:
            school_year = int(school_year)
        else:
            now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9)))
            school_year = now.year

        if out_file:
            pass
        else:
            out_file = f"calendar_sy{school_year}.xlsx"

        st = datetime.datetime(
            year=school_year,
            month=4,
            day=1,
            tzinfo=datetime.timezone(datetime.timedelta(hours=9)),
        )
        ed = datetime.datetime(
            year=school_year + 1,
            month=3,
            day=31,
            tzinfo=datetime.timezone(datetime.timedelta(hours=9)),
        )
        page_token = None
        while True:
            events_result = (
                service.events()
                .list(
                    calendarId=CAL_ID,
                    timeMin=st.isoformat(),
                    timeMax=ed.isoformat(),
                    maxResults=1000,
                    singleEvents=True,
                    orderBy="startTime",
                    pageToken=page_token,
                )
                .execute()
            )
            write_header(ws)
            events = events_result.get("items", [])
            for event in events:
                # ic(event)
                counter += 1
                write_row(ws, counter, event)
            page_token = events_result.get("nextPageToken")
            if not page_token:
                break

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 16
        ws.column_dimensions["G"].width = 30

        # 列選択で保護解除できないので力技
        for i in range(1, 1000):
            ws.cell(i, 1).protection = Protection(locked=False)
            ws.cell(i, 2).protection = Protection(locked=False)
            ws.cell(i, 3).protection = Protection(locked=False)
            ws.cell(i, 4).protection = Protection(locked=False)

        ws.freeze_panes = "B2"
        # ws.protection.password = "hinogaku"
        ws.protection.enable()
        ws.protection.objects = False
        ws.protection.scenarios = False
        ws.protection.insertRows = False
        ws.protection.deleteRows = False
        ws.protection.sort = False
        ws.protection.formatCells = False
        ws.protection.formatRows = False
        ws.protection.formatCells = False
        ws.protection.formatColumns = False
        ws.protection.formatRows = False
        wb.save(out_file)
        ic(out_file)

        return True

    except Exception as e:
        ic(e)


def upload_events(service, school_year, in_file):
    """エクセルの内容をカレンダーへ反映する"""

    if school_year:
        school_year = int(school_year)
    else:
        now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9)))
        school_year = now.year

    if in_file:
        pass
    else:
        in_file = f"calendar_sy{school_year}.xlsx"

    ic(in_file)
    wb = openpyxl.load_workbook(in_file)
    ws = wb.active  # wb["Sheet1"]
    max_row = ws.max_row
    counter = 0
    for row in ws:
        counter += 1
        # if counter > 10:  # debug
        #     break
        if row[0].value == "開始日":
            # skip
            print(f"{counter}/{max_row}: skip title row")
            pass
        else:
            g_event = create_event_from_row(row)
            if g_event:
                event = update_event(service, g_event)
                if event and event.get("summary") != "":
                    print(
                        f"{counter}/{max_row}: Updated. {event["start"].get("dateTime",event["start"].get("date"))} - {event["end"].get("dateTime",event["end"].get("date"))}: {event["summary"]} {event.get("description")}"
                    )
                    # write back to excel sheet
                    ws.cell(counter, 7, g_event["id"])
                    created = iso2jst(event.get("created"))
                    updated = iso2jst(event.get("updated"))
                    ws.cell(counter, 5, created.replace(tzinfo=None))
                    ws.cell(counter, 5).number_format = "yyyy/mm/dd hh:mm"
                    ws.cell(counter, 6, updated.replace(tzinfo=None))
                    ws.cell(counter, 6).number_format = "yyyy/mm/dd hh:mm"
                elif event and event.get("summary") == "":
                    print(f"{counter}/{max_row}: Deleted. {g_event["id"]}")
                else:
                    print(
                        f"{counter}/{max_row}: Skipped. {g_event["start"].get("dateTime",g_event["start"].get("date"))} - {g_event["end"].get("dateTime",g_event["end"].get("date"))}: {g_event.get("summary")}"
                    )
            else:
                # ignore None
                # print(f"{counter}/{max_row}: ignore")
                pass
    wb.save(in_file)

    try:
        pass

        # event = update_event(service, event)
        # print ('Event created: %s' % (event.get('htmlLink')))

    except HttpError as error:
        print(f"An error occurred: {error}")


def list_calendar(service):
    """カレンダー一覧を取得する"""
    try:
        page_token = None
        while True:
            calendar_list = service.calendarList().list(pageToken=page_token).execute()
            for calendar_list_entry in calendar_list["items"]:
                ic(calendar_list_entry)
                # print(calendar_list_entry["summary"])
                # print(calendar_list_entry)
            page_token = calendar_list.get("nextPageToken")
            if not page_token:
                break
    except HttpError as error:
        print(f"An error occurred: {error}")


def create_event_from_row(row):
    """カレンダーイベントを作成する"""
    summary = row[2].value
    if summary == None or summary == "":
        # 削除レコード確認
        if (
            (row[0].value == None or row[0].value == "" or row[0].value == 0)
            and row[6].value != None
            and row[6].value != ""
        ):
            # 削除
            event = {"id": f"{row[6].value}", "summary": ""}
            return event
        else:
            # 不正なレコード。読み飛ばし対象
            return None

    description = row[3].value or ""

    if len(row) > 6:
        event_id = row[6].value
    else:
        event_id = None
    if event_id == None or event_id == "" or event_id == 0:
        print("new record found.", row[2].value)
        event_id = uuid.uuid4().hex

    event = {
        "id": f"{event_id}",
        "summary": f"{summary}",
        "start": {},
        "end": {},
        "description": f"{description}",
        "reminders": {"useDefault": False},
    }

    start = row[0].value
    end = row[1].value

    # validation
    if type(start) == datetime.date or type(start) == datetime.datetime:
        # OK
        pass
    else:
        # NG
        return None
    if type(end) == datetime.date or type(end) == datetime.datetime:
        # OK
        pass
    else:
        # 日付省略パターン(まずは"その日まで"として設定)
        end = start

    event_type = ""

    # date or datetime 判定
    if (
        start.strftime("%H:%M:%S") == "00:00:00"
        and end.strftime("%H:%M:%S") == "00:00:00"
    ):
        event_type = "date"
    else:
        event_type = "datetime"

    if event_type == "date":
        event["start"]["date"] = start.strftime("%Y-%m-%d")
        end = end + datetime.timedelta(days=1)
        event["end"]["date"] = end.strftime("%Y-%m-%d")
    else:
        # event_type == datetime
        start_iso = start.replace(
            tzinfo=datetime.timezone(datetime.timedelta(hours=+9))
        ).isoformat()
        event["start"]["dateTime"] = start_iso
        event["start"]["timeZone"] = "Asia/Tokyo"
        # ↓時間指定されている場合は日付を+1日する必要はない
        # end = start + datetime.timedelta(days=1)
        end_iso = end.replace(
            tzinfo=datetime.timezone(datetime.timedelta(hours=+9))
        ).isoformat()
        event["end"]["dateTime"] = end_iso
        event["end"]["timeZone"] = "Asia/Tokyo"

    return event
