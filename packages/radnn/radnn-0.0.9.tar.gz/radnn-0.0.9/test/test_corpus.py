from radnn import mlsys, FileSystem
from openpyxl import load_workbook
from datasets import TextQuestionSample, CytaChatbotDataset
from chatgpt import ChatGPTAPI
import re
from tqdm import tqdm

def split_number(text):
    match = re.match(r'^(\d+[.)])\s*(.*)', text)
    if match:
        number = match.group(1)  # The numeric part with ) or .
        rest = match.group(2)    # The remaining text
        return number, rest
    else:
        return None, text

mlsys.filesys = FileSystem()
oDataset = CytaChatbotDataset(mlsys.filesys.datasets.subfs("CYTACHATBOT"))


API_KEY = "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"
oAPI = ChatGPTAPI(API_KEY)

for sLang in ["EL", "EN"]:
    if not oDataset.load_question_answers(sLang):
        sImportFileName = oDataset.fs.subfs("source").file(f"CytaChatbot_v2-{sLang}.xlsx")
        # Load workbook and select a sheet
        wb = load_workbook(sImportFileName)
        sheet = wb.active  # Or wb['SheetName']

        # Iterate over rows

        bIsQuestion = True
        nIndex = -1
        for row in sheet.iter_rows(values_only=True):
            assert row[1] is None, "More columns"
            bIsQuestion = row[0] is not None
            bIsAnswer = row[0]  is not None
            if bIsQuestion:
                nID, sQuestion = split_number(row[0].strip())
                bIsQuestion =  nID is not None
            if bIsQuestion:
                nIndex += 1
                oSample = TextQuestionSample()
                oDataset.append(oSample)

                sRow = row[0].strip()
                nID = None
                if (sRow[1] == ".") or (sRow[1] == ")"):
                    nID = int(sRow[:1])
                    sQuestion = sRow[2:]
                if (sRow[2] == ".") or (sRow[2] == ")"):
                    nID = int(sRow[:2])
                    sQuestion = sRow[3:]
                if (sRow[3] == ".") or (sRow[3] == ")"):
                    nID = int(sRow[:3])
                    sQuestion = sRow[4:]
                oSample.id = nID

                oSample.question = sQuestion.strip()
                oSample.answer = ""
                bIsQuestion = False
            elif bIsAnswer:
                sRow = row[0].strip()
                if oSample.answer == "":
                    oSample.answer += "\n" + sRow
                else:
                    oSample.answer = sRow

                print(oSample)


        oDataset.save_question_answers(sLang)

    EXTRA_QUESTIONS_COUNT = 30
    for nIndex, oSample in tqdm(enumerate(oDataset)):
        try:
          if sLang == "EL":
            prompt = f"Γράψε {EXTRA_QUESTIONS_COUNT} παραλλαγές της ερώτησης: '{oSample.question}', στην απάντηση βάλε μόνο τις παραλλαγές"
          else:
            prompt = f"Write {EXTRA_QUESTIONS_COUNT} variations of the question: '{oSample.question}', keep only the variations in the answer"
          response = oAPI.generate(prompt)
          oSample.question_alt = response.split('\n')

          print(oSample)
          print(oSample.question_alt)
        except Exception as e:
          new_questions = f"Error: {e}"

        oDataset.save_question_answers_aug(sLang)

