import random

from radnn import mlsys, FileSystem
from openpyxl import load_workbook
from datasets import TextQuestionSample, CytaChatbotDataset, StratifiedCytaChatbotDataset
from chatgpt import ChatGPTAPI
import re
from tqdm import tqdm
import ast
import numpy as np

LANG = "EL"

def split_alt_id_question(question):
    result = []
    match = re.match(r"(\d+)\.\s*(.+)", question)
    if match:
        result.append(match.group(1))
        result.append(match.group(2))
    return result


API_KEY = "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"
oAPI = ChatGPTAPI(API_KEY)



mlsys.filesys = FileSystem()
oDataset = CytaChatbotDataset(mlsys.filesys.datasets.subfs("CYTACHATBOT"))
if not oDataset.load_question_answers_aug(LANG):
    print("[>] Attaching alternative questions ...")
    oDataset.load_question_answers(LANG)


    # Calculate the difference between consecutive elements
    oIds = np.asarray([oSample.id for oSample in oDataset], dtype=np.int32)
    diffs = np.diff(oIds)

    # Find indices where the difference is more than 1 (indicates a gap)
    gap_indices = np.where(diffs > 1)[0]
    if len(gap_indices) > 0:
        print("[x] Gaps in sample numbering:")
        # Report the missing numbers
        for index in gap_indices:
            start = oIds[index] + 1
            end = oIds[index + 1] - 1
            missing = list(range(start, end + 1))
            print(f"|__ Missing: {missing}")

    if oDataset.fs.subfs("source").exists(f"CytaChatbot_v1_Targets-{LANG}.xlsx"):
        print("|__ Attaching annotations")
        sImportFileName = oDataset.fs.subfs("source").file(f"CytaChatbot_v1_Targets-{LANG}.xlsx")
        oWorkbook = load_workbook(sImportFileName)
        oSheet = oWorkbook.active  # Or wb['SheetName']
        oTargets = []
        for nIndex, oRow in enumerate(oSheet.iter_rows(values_only=True)):
            if (nIndex > 0) and (oRow[0] is not None):
                try:
                    nOnehot = np.asarray(oRow[2:5], dtype=np.int32)
                except Exception as e:
                    print(nIndex, oRow)
                    raise
                assert nOnehot.sum() == 1, "More than one tags"

                nID = nIndex - 1
                sID = oRow[0]
                sSample = oRow[1].strip().replace("?", ";")
                sSampleV2 = oDataset[nID].question.strip().replace("?", ";")
                nTarget = np.argmax(nOnehot)
                oTargets.append(nTarget)
                oDataset[nID].annotations = nTarget
                #if sSample != sSampleV2:
                #    print(nSampleID, sSample, "!=", sSampleV2, nTarget)
                #else:
                    #print(nSampleID, sSample, nTarget)
    else:
        # TODO: EN Annotations
        random.seed(2025)
        oTargets = []
        for nIndex, oTextRecord in enumerate(oDataset):
            nTarget = random.randint(0, 2)
            oTargets.append(nTarget)
            oTextRecord.annotations = nTarget


    print("|__ Adding augmentations")
    if False:
        EXTRA_QUESTIONS_COUNT = 30
        for nIndex, oTextRecord in enumerate(tqdm(oDataset)):
            try:
              if LANG == "EL":
                prompt = f"Γράψε {EXTRA_QUESTIONS_COUNT} παραλλαγές της ερώτησης: '{oTextRecord.question}', στην απάντηση βάλε μόνο τις παραλλαγές"
              else:
                prompt = f"Write {EXTRA_QUESTIONS_COUNT} variations of the question: '{oTextRecord.question}', keep only the variations in the answer"
              response = oAPI.generate(prompt)
              oTextRecord.question_alt = response.split('\n')

              print(oTextRecord)
              print(oTextRecord.question_alt)
            except Exception as e:
              new_questions = f"Error: {e}"

            oDataset.save_question_answers_aug(LANG)
    else:
        mlsys.filesys = FileSystem()
        oFS = mlsys.filesys.datasets.subfs("CYTACHATBOT")
        sFileContents = oFS.text.load(f"Augmented_{LANG}.txt")
        oFilteredLines = []
        for sLine in sFileContents.splitlines():
            if sLine.startswith("['") or sLine.startswith('["'):
                oFilteredLines.append(sLine)

        nSampleIndex = 0
        for sLine in oFilteredLines:
            oList = ast.literal_eval(sLine)
            oListClean = []
            for x in oList:
                oParts = split_alt_id_question(x)
                if len(oParts) > 1:
                    oListClean.append(oParts[1].strip().replace("?", ";"))
            #print(sID, oListClean)
            assert len(oListClean) == 30, "Wrong count of alternative questions"
            oDataset[nSampleIndex].question_alt = oListClean
            oDataset[nSampleIndex].question = oDataset[nSampleIndex].question.strip().replace("?", ";")
            nSampleIndex += 1

    oNewDataSet = CytaChatbotDataset(mlsys.filesys.datasets.subfs("CYTACHATBOT"))
    for oSample in oDataset:
        oNewDataSet.append(oSample)
    oNewDataSet.save_question_answers_aug(LANG)
else:
    oNewDataSet = oDataset
    oTargets = [oSample.annotations for oSample in oDataset]

nClassHistogram, bin_edges = np.histogram(oTargets, bins=3)





def StratifiedBalancing(dataset_augmented, class_histogram=[89, 157, 358]):
    nClassHistogram = class_histogram
    nClasses = len(nClassHistogram)
    nExtraSamples = 30
    nMinorityClassCount = np.min(nClassHistogram)
    nMaxSamples = nMinorityClassCount*(nExtraSamples + 1)
    nTarget = (nMaxSamples // 50) * 50

    nMaxSamplesPerClass   = np.zeros(nClasses, np.int32)
    nTargetClassHistogram = np.zeros(nClasses, np.int32)
    nExtraSamplesPerClass = np.zeros(nClasses, np.int32)
    nMinusOneSamplesCount = np.zeros(nClasses, np.int32)

    for nIndex, nOriginalCount in enumerate(nClassHistogram):
        if nOriginalCount != nMinorityClassCount:
            nExtraSamplesCeil = int(np.ceil(nTarget / nOriginalCount)) - 1
            nMaxSamplesPerClass[nIndex]   = nOriginalCount * (nExtraSamplesCeil + 1)
            nExtraSamplesPerClass[nIndex] = nExtraSamplesCeil
            nTargetClassHistogram[nIndex] = nTarget
            nMinusOneSamplesCount[nIndex] = nTarget - nMaxSamplesPerClass[nIndex]
        else:
            nMaxSamplesPerClass[nIndex] = nMaxSamples
            nExtraSamplesPerClass[nIndex] = nExtraSamples
            nTargetClassHistogram[nIndex] = nTarget
            nMinusOneSamplesCount[nIndex] = nTarget - nMaxSamples



    print(nClassHistogram)
    print(nExtraSamplesPerClass)
    print(nMaxSamplesPerClass)
    print(nMinusOneSamplesCount)
    print(nTargetClassHistogram)

    nMinusOneLimit = nClassHistogram + nMinusOneSamplesCount
    nClassOccurences = np.zeros(nClasses, np.int32)

    oStratifiedDS = StratifiedCytaChatbotDataset(dataset_augmented.fs)

    for oSample in dataset_augmented:
        sAnswer = oSample.answer
        nClassIndex = oSample.annotations
        nClassOccurences[nClassIndex] += 1
        nExtraSamples = nExtraSamplesPerClass[nClassIndex]
        if nClassOccurences[nClassIndex] > nMinusOneLimit[nClassIndex]:
            nExtraSamples -= 1

        nBaseNewID = 1000000 + (oSample.id*1000)
        oNewSample = TextQuestionSample(nBaseNewID, oSample.question, sAnswer, nClassIndex)
        oQuestionSamples = list()
        oStratifiedDS.class_questions[nClassIndex][oSample.id] = oQuestionSamples
        oQuestionSamples.append(oNewSample)
        for nExtraSampleIndex in range(nExtraSamples):
            oNewSample = TextQuestionSample(nBaseNewID + 1 + nExtraSampleIndex, oSample.question_alt[nExtraSampleIndex], sAnswer, nClassIndex)
            oQuestionSamples.append(oNewSample)


    return oStratifiedDS



oStratifiedDS = StratifiedBalancing(oNewDataSet, nClassHistogram)
oStratifiedDS.save_questions(LANG)

if False:
    for k, v in oStratifiedDS.class_questions[2].items():
        for s in v:
            print(s.question)

